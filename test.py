from openpyxl import Workbook, load_workbook

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb_input = load_workbook("C:/Users/KJH/Desktop/납품계획서.xlsx", data_only=True)
load_wb_output = load_workbook("C:/Users/KJH/Desktop/Output.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

# 셀 주소로 값 출력
print(load_ws['B2'].value)

# 셀 좌표로 값 출력
print(load_ws.cell(3, 2).value)

# 지정한 셀의 값 출력

get_cells = load_ws['B3': 'B6']
for row in get_cells:
    for cell in row:
        print(cell.value)

# 모든 행 단위로 출력

for row in load_ws.rows:
    print(row)

# 모든 열 단위로 출력

for column in load_ws.columns:
    print(column)

# 모든 행과 열 출력

all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)

load_ws.cell(3, 3, 51470)
load_ws.cell(4, 3, 21470)
load_ws.cell(5, 3, 1470)
load_ws.cell(6, 3, 6470)
load_wb.save("C:/Users/Administrator/Desktop/기준/프로그래밍/과제대행/주식데이터크롤링/output.xlsx")

# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '숫자'

# 행 단위로 추가
write_ws.append([1, 2, 3])

# 셀 단위로 추가
write_ws.cell(5, 5, '5행5열')
write_wb.save("C:/Users/KJH/Desktop/test.xlsx")